#!C:\Python38\python.exe

from os.path import isfile
from os.path import abspath
from ntpath import basename
from openpyxl import load_workbook
from sys import argv, exit
from win32com.client import Dispatch
import win32com.client
import shutil
import os
import sys
import re
import json
from Global_Properties import Global_Properties
jobProperties = Global_Properties()
jsonData = jobProperties.targetCantataDir
with open(jsonData, 'r') as f:
    data = json.load(f)
global glbCantataWS,glbMergeWS
glbCantataWS = data["Cantata"]["Workspace"]
rootReport = f'{glbCantataWS}/{jobProperties.moduleRunning.capitalize()}_TestLog/Env'
glbMergeWS = f'{rootReport}/{jobProperties.moduleRunning.capitalize()}'
gReportFile = f'{glbMergeWS}/report/Coverage_Report_Template.xlsm'
gTargetList = f'{glbMergeWS}/target.txt'

def RunReportMacro(Report):
    excel = Excel_Application()
    excel.DisplayAlerts = False
    excel.Workbooks.Open(Report, ReadOnly=False)
    excel.WindowState = win32com.client.constants.xlMaximized
    MacroName = 'Sheet17.Sample'
    print('[+] Running macro: ' + MacroName)
    print('[+] Parsing merged file to report.')
    excel.Application.Run(MacroName)
    excel.ActiveWorkbook.Save()
    excel.ActiveWorkbook.Close()


def Excel_Application():
    try:
        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    except AttributeError:
        # Corner case dependencies.
        # Remove cache and try again.
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get(
            'LOCALAPPDATA'), 'Temp', 'gen_py'))
        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
    return excel


def InsertTargetFile(WbReport, TargetFileList):
    print('[+] Inserting target source file names to report.')
    SheetReport = WbReport['Result Summary']
    for i in range(len(TargetFileList)):
        SheetReport.cell(column=3, row=9+i, value=TargetFileList[i])
    WbReport.save(gReportFile)
    WbReport.close()


def main():
    # Check report file location
    if isfile(gReportFile):
        print('[+] Loading report excel file.')
        WbReport = load_workbook(filename=gReportFile,
                                 read_only=False, keep_vba=True)
    else:
        print('[+] ERROR: Report file is not exist!!!')
        exit(1)
    # Get source file list from target file.
    ListFile = []
    if isfile(gTargetList):
        with open(gTargetList) as file:
            ListFile = file.read().splitlines()
        print('[+] List of file:')
        for i in range(len(ListFile)):
            ListFile[i] = ListFile[i] + '.c'
            print('[+] ' + ListFile[i])
    else:
        print('[+] ERROR: List of source files is not exist')
        exit(1)
    # Insert target file to report file.
    InsertTargetFile(WbReport, ListFile)
    # Run macro
    RunReportMacro(gReportFile)
    print('[+] Done!!!')


if __name__ == "__main__":
    main()
